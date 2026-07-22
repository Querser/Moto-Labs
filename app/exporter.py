"""Simple one-row-per-lap Excel export."""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from datetime import timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import LapRecord, Race

NS_PER_DAY = 86_400_000_000_000
DURATION_FORMAT = "[m]:ss.000"
UTC_FORMAT = "yyyy-mm-dd hh:mm:ss.000"


@dataclass(frozen=True, slots=True)
class ExportedWorkbook:
    path: Path
    filename: str


def safe_filename(value: str) -> str:
    value = unicodedata.normalize("NFKC", value)
    value = re.sub(r"[\\/:*?\"<>|\x00-\x1f]", "_", value)
    value = re.sub(r"\s+", "_", value.strip(" ._"))
    return (value[:80] or "race").strip(" ._") or "race"


def excel_text(value: str | None) -> str | None:
    if value is None:
        return None
    value = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", value)
    return "'" + value if value.startswith(("=", "+", "-", "@")) else value


class ExcelLapExporter:
    def __init__(self, output_dir: str | Path) -> None:
        self.output_dir = Path(output_dir).expanduser().resolve()

    def export(self, session: Session, race_id: int) -> ExportedWorkbook:
        race = session.get(Race, race_id)
        if race is None:
            raise LookupError(f"race {race_id} was not found")
        laps = list(
            session.scalars(
                select(LapRecord)
                .where(LapRecord.race_id == race_id)
                .order_by(LapRecord.detected_at_utc, LapRecord.id)
            )
        )
        self.output_dir.mkdir(parents=True, exist_ok=True)
        stamp = (race.finished_at_utc or race.created_at).strftime("%Y%m%d_%H%M%S")
        filename = f"{safe_filename(race.name)}_{stamp}.xlsx"
        path = (self.output_dir / filename).resolve()
        if path.parent != self.output_dir:
            raise ValueError("unsafe export path")

        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = "Laps"
        metadata: list[tuple[str, Any]] = [
            ("Race", excel_text(race.name)),
            ("Description", excel_text(race.description)),
            ("Required Laps", race.required_laps),
            ("Started At (UTC)", _excel_datetime(race.started_at_utc)),
            ("Finished At (UTC)", _excel_datetime(race.finished_at_utc)),
        ]
        for row in metadata:
            sheet.append(row)
        sheet.append([])
        header_row = sheet.max_row + 1
        headers = [
            "Motorcycle Number",
            "Lap",
            "Lap Time",
            "Race Elapsed Time",
            "Recorded At",
        ]
        sheet.append(headers)
        for cell in sheet[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="176B87")
        for lap in laps:
            sheet.append(
                [
                    excel_text(lap.racing_number),
                    lap.lap_number,
                    lap.lap_time_ns / NS_PER_DAY,
                    lap.race_elapsed_ns / NS_PER_DAY,
                    _excel_datetime(lap.detected_at_utc),
                ]
            )
            current = sheet.max_row
            sheet.cell(current, 1).number_format = "@"
            sheet.cell(current, 3).number_format = DURATION_FORMAT
            sheet.cell(current, 4).number_format = DURATION_FORMAT
            sheet.cell(current, 5).number_format = UTC_FORMAT
        sheet.freeze_panes = f"A{header_row + 1}"
        sheet.auto_filter.ref = f"A{header_row}:E{max(header_row, sheet.max_row)}"
        for cells in sheet.columns:
            length = max(
                (len(str(cell.value)) for cell in cells if cell.value is not None),
                default=0,
            )
            column = cells[0].column
            assert isinstance(column, int)
            sheet.column_dimensions[get_column_letter(column)].width = min(
                max(12, length + 2), 42
            )
        workbook.save(path)
        return ExportedWorkbook(path, filename)


def _excel_datetime(value: Any) -> Any:
    if value is None:
        return None
    if value.tzinfo is None:
        return value
    return value.astimezone(timezone.utc).replace(tzinfo=None)


__all__ = ["ExcelLapExporter", "ExportedWorkbook", "excel_text", "safe_filename"]
