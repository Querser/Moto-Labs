from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.domain import LapTimingService, PassageCandidate
from app.exporter import DURATION_FORMAT, ExcelLapExporter


def test_export_has_one_row_per_lap_and_keeps_007(session: Session, tmp_path: Path) -> None:
    service = LapTimingService(session)
    race = service.create_race(name="../Race", required_laps=2, camera_identifier="synthetic")
    service.start(race.id, 1_000)
    for index, at_ns in enumerate((2_000, 4_000), start=1):
        service.record_passage(
            race.id,
            PassageCandidate(
                "007", at_ns, datetime.now(timezone.utc), 0.99, idempotency_key=str(index)
            ),
        )
    session.commit()

    exported = ExcelLapExporter(tmp_path).export(session, race.id)
    assert exported.path.parent == tmp_path.resolve()
    workbook = load_workbook(exported.path)
    sheet = workbook["Laps"]
    header = next(
        row
        for row in range(1, sheet.max_row + 1)
        if sheet.cell(row, 1).value == "Motorcycle Number"
    )
    assert [sheet.cell(header + offset, 1).value for offset in (1, 2)] == ["007", "007"]
    assert [sheet.cell(header + offset, 2).value for offset in (1, 2)] == [1, 2]
    assert sheet.cell(header + 1, 1).number_format == "@"
    assert sheet.cell(header + 1, 3).number_format == DURATION_FORMAT
